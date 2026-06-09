"""
Genera el reporte final en Excel con:
- Hoja 1: Resumen ejecutivo
- Hoja 2: Datos crudos por municipalidad
- Hoja 3: Estadística descriptiva (rendimiento)
- Hoja 4: Estadística descriptiva (frescura/transparencia)
- Hoja 5: Estadística descriptiva (seguridad)
- Hoja 6: Proporciones de cumplimiento
- Hoja 7: Comparación entre departamentos (Kruskal-Wallis)
- Hoja 8: Correlaciones (Spearman)

Adicionalmente genera gráficas PNG en data/reports/graficas/ y un CSV plano.
"""
from __future__ import annotations

from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import REPORTS_DIR
from src.analysis import stats as A
from src.logger import get_logger

log = get_logger(__name__)

sns.set_style("whitegrid")
plt.rcParams["font.family"] = "DejaVu Sans"


# Columnas relevantes para análisis numérico
COLS_PERF = [
    "ttfb_ms", "tiempo_total_ms", "tamanio_kb",
    "imgs_pct_con_alt", "score_local_performance",
    "pagespeed_score_performance", "pagespeed_score_accessibility",
    "pagespeed_lcp_ms", "pagespeed_cls",
]

COLS_FRESHNESS = [
    "snapshots_total", "snapshots_unicos",
    "intervalo_medio_dias", "intervalo_mediana_dias",
    "dias_desde_ultima_actualizacion",
    "actualizaciones_unicas_2025_2026",
    "laip_pct_cumplimiento", "score_local_freshness",
]

COLS_SECURITY = [
    "ssl_dias_restantes", "headers_seguridad_pct",
    "score_local_security",
]

COLS_BOOL_PERF = [
    "reachable", "https", "tiene_viewport", "tiene_lang",
    "tiene_charset", "tiene_h1", "tiene_favicon",
]

COLS_BOOL_FRESHNESS = [
    "laip_transparencia", "laip_presupuesto", "laip_compras",
    "laip_personal", "laip_servicios", "laip_estructura", "laip_contacto",
]

COLS_BOOL_SECURITY = [
    "ssl_ok", "tls_aceptable", "cert_vigencia_suficiente",
    "redirige_a_https",
    "header_hsts", "header_csp", "header_x_frame_options",
    "header_x_content_type_options", "header_referrer_policy",
    "header_permissions_policy",
    "expone_server_header", "expone_x_powered_by",
]


# ---------- Estilos ----------

FONT_HEADER = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FILL_HEADER = PatternFill("solid", start_color="1F4E78")
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="1F4E78")
FONT_BODY = Font(name="Arial", size=10)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
BORDER_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _formatear_hoja(ws, df: pd.DataFrame, fila_inicio: int = 1) -> None:
    """Aplica formato a una hoja con encabezados y bordes."""
    # Encabezados
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=fila_inicio, column=j, value=str(col))
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN

    # Datos
    for i, fila in enumerate(df.itertuples(index=False), start=fila_inicio + 1):
        for j, val in enumerate(fila, start=1):
            c = ws.cell(row=i, column=j, value=_valor_safe(val))
            c.font = FONT_BODY
            c.alignment = ALIGN_LEFT
            c.border = BORDER_THIN

    # Ancho de columnas
    for j, col in enumerate(df.columns, start=1):
        ancho = max(12, min(40, len(str(col)) + 4))
        ws.column_dimensions[get_column_letter(j)].width = ancho

    ws.freeze_panes = ws.cell(row=fila_inicio + 1, column=1)


def _valor_safe(v):
    """Convierte valores a tipos compatibles con openpyxl."""
    if isinstance(v, (list, dict, set, tuple)):
        return str(v)
    if pd.isna(v):
        return None
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return float(v) if not np.isnan(v) else None
    if isinstance(v, (np.bool_,)):
        return bool(v)
    return v


# ---------- Gráficas ----------

def _generar_graficas(df: pd.DataFrame, graficas_dir: Path) -> List[Path]:
    """Genera PNGs con visualizaciones clave."""
    graficas_dir.mkdir(parents=True, exist_ok=True)
    paths = []

    # 1. Distribución de scores agregados
    try:
        fig, axes = plt.subplots(1, 3, figsize=(15, 4))
        for ax, col, titulo in zip(
            axes,
            ["score_local_performance", "score_local_freshness", "score_local_security"],
            ["Rendimiento (OE1)", "Frescura/LAIP (OE2)", "Seguridad (OE3)"],
        ):
            if col in df.columns:
                s = pd.to_numeric(df[col], errors="coerce").dropna()
                if not s.empty:
                    ax.hist(s, bins=10, edgecolor="black", color="#1F4E78")
                    ax.set_title(titulo)
                    ax.set_xlabel("Score 0-100")
                    ax.set_ylabel("Municipalidades")
                    ax.axvline(s.mean(), color="red", linestyle="--",
                               label=f"Media={s.mean():.1f}")
                    ax.legend()
        plt.tight_layout()
        p = graficas_dir / "01_distribucion_scores.png"
        plt.savefig(p, dpi=120, bbox_inches="tight")
        plt.close()
        paths.append(p)
    except Exception as ex:
        log.warning("No se pudo generar gráfica scores: %s", ex)

    # 2. Tiempo de carga por departamento
    try:
        if "departamento" in df.columns and "tiempo_total_ms" in df.columns:
            fig, ax = plt.subplots(figsize=(10, 5))
            sub = df.dropna(subset=["tiempo_total_ms"]).copy()
            sub["tiempo_s"] = sub["tiempo_total_ms"] / 1000
            sns.boxplot(data=sub, x="departamento", y="tiempo_s", ax=ax,
                        color="#5B9BD5")
            ax.set_title("Tiempo de carga total por departamento (s)")
            ax.set_xlabel("Departamento")
            ax.set_ylabel("Segundos")
            plt.xticks(rotation=30, ha="right")
            plt.tight_layout()
            p = graficas_dir / "02_carga_por_departamento.png"
            plt.savefig(p, dpi=120, bbox_inches="tight")
            plt.close()
            paths.append(p)
    except Exception as ex:
        log.warning("No se pudo generar gráfica carga: %s", ex)

    # 3. Cumplimiento LAIP por categoría
    try:
        laip_cols = [c for c in df.columns if c.startswith("laip_") and c.endswith(
            ("transparencia", "presupuesto", "compras", "personal",
             "servicios", "estructura", "contacto")
        )]
        if laip_cols:
            pcts = {}
            for c in laip_cols:
                v = df[c].fillna(False).astype(bool)
                pcts[c.replace("laip_", "")] = v.mean() * 100
            fig, ax = plt.subplots(figsize=(9, 5))
            categorias = list(pcts.keys())
            valores = list(pcts.values())
            ax.barh(categorias, valores, color="#70AD47")
            ax.set_xlabel("% de portales que la presentan")
            ax.set_title("Cumplimiento de indicadores LAIP (Decreto 57-2008)")
            ax.set_xlim(0, 100)
            for i, v in enumerate(valores):
                ax.text(v + 1, i, f"{v:.1f}%", va="center")
            plt.tight_layout()
            p = graficas_dir / "03_laip_cumplimiento.png"
            plt.savefig(p, dpi=120, bbox_inches="tight")
            plt.close()
            paths.append(p)
    except Exception as ex:
        log.warning("No se pudo generar gráfica LAIP: %s", ex)

    # 4. Headers de seguridad presentes
    try:
        header_cols = [
            "header_hsts", "header_csp", "header_x_frame_options",
            "header_x_content_type_options", "header_referrer_policy",
            "header_permissions_policy",
        ]
        header_cols = [c for c in header_cols if c in df.columns]
        if header_cols:
            pcts = {}
            for c in header_cols:
                pcts[c.replace("header_", "").upper()] = (
                    df[c].fillna(False).astype(bool).mean() * 100
                )
            fig, ax = plt.subplots(figsize=(9, 5))
            categorias = list(pcts.keys())
            valores = list(pcts.values())
            ax.barh(categorias, valores, color="#C00000")
            ax.set_xlabel("% de portales que lo configuran")
            ax.set_title("Headers de seguridad HTTP presentes")
            ax.set_xlim(0, 100)
            for i, v in enumerate(valores):
                ax.text(v + 1, i, f"{v:.1f}%", va="center")
            plt.tight_layout()
            p = graficas_dir / "04_security_headers.png"
            plt.savefig(p, dpi=120, bbox_inches="tight")
            plt.close()
            paths.append(p)
    except Exception as ex:
        log.warning("No se pudo generar gráfica security headers: %s", ex)

    # 5. Snapshots Wayback por año (agregado)
    try:
        if "snapshots_por_anio" in df.columns:
            por_anio_agg: Dict[str, int] = {}
            for v in df["snapshots_por_anio"].dropna():
                if isinstance(v, dict):
                    for anio, n in v.items():
                        por_anio_agg[str(anio)] = por_anio_agg.get(str(anio), 0) + int(n)
                elif isinstance(v, str):
                    try:
                        import ast
                        d = ast.literal_eval(v)
                        if isinstance(d, dict):
                            for anio, n in d.items():
                                por_anio_agg[str(anio)] = por_anio_agg.get(str(anio), 0) + int(n)
                    except Exception:
                        pass
            if por_anio_agg:
                anios = sorted(por_anio_agg.keys())
                valores = [por_anio_agg[a] for a in anios]
                fig, ax = plt.subplots(figsize=(9, 4))
                ax.bar(anios, valores, color="#7030A0")
                ax.set_title("Snapshots únicos Wayback Machine — agregado regional")
                ax.set_xlabel("Año")
                ax.set_ylabel("Snapshots")
                for i, v in enumerate(valores):
                    ax.text(i, v + 0.5, str(v), ha="center")
                plt.tight_layout()
                p = graficas_dir / "05_wayback_por_anio.png"
                plt.savefig(p, dpi=120, bbox_inches="tight")
                plt.close()
                paths.append(p)
    except Exception as ex:
        log.warning("No se pudo generar gráfica wayback: %s", ex)

    return paths


# ---------- Resumen ejecutivo ----------

def _resumen_ejecutivo(df: pd.DataFrame) -> pd.DataFrame:
    """Tabla con los KPIs principales del estudio."""
    total = len(df)
    reach = int(df["reachable"].fillna(False).astype(bool).sum()) if "reachable" in df.columns else 0

    def pct_true(col):
        if col not in df.columns:
            return None
        return round(df[col].fillna(False).astype(bool).mean() * 100, 2)

    def media(col):
        if col not in df.columns:
            return None
        s = pd.to_numeric(df[col], errors="coerce").dropna()
        return round(s.mean(), 2) if not s.empty else None

    filas = [
        ("Municipalidades evaluadas (total)", total, ""),
        ("Municipalidades alcanzables (reachable)", reach, f"{round(reach/total*100,2) if total else 0}%"),
        ("", "", ""),
        ("— OE1: RENDIMIENTO Y ACCESIBILIDAD —", "", ""),
        ("Tiempo medio de carga total (ms)", media("tiempo_total_ms"), ""),
        ("TTFB medio (ms)", media("ttfb_ms"), ""),
        ("Peso medio de página (KB)", media("tamanio_kb"), ""),
        ("% portales con viewport móvil", pct_true("tiene_viewport"), "%"),
        ("% portales con HTTPS", pct_true("https"), "%"),
        ("% portales con atributo lang", pct_true("tiene_lang"), "%"),
        ("% imágenes con alt (media de portales)", media("imgs_pct_con_alt"), "%"),
        ("Score local rendimiento medio", media("score_local_performance"), "/100"),
        ("", "", ""),
        ("— OE2: FRESCURA Y TRANSPARENCIA LAIP —", "", ""),
        ("Snapshots Wayback medios por portal", media("snapshots_unicos"), ""),
        ("Días promedio desde última actualización", media("dias_desde_ultima_actualizacion"), ""),
        ("Cumplimiento LAIP medio", media("laip_pct_cumplimiento"), "%"),
        ("% portales con sección de transparencia", pct_true("laip_transparencia"), "%"),
        ("% portales con sección de presupuesto", pct_true("laip_presupuesto"), "%"),
        ("% portales con sección de compras (Guatecompras)", pct_true("laip_compras"), "%"),
        ("Score local frescura medio", media("score_local_freshness"), "/100"),
        ("", "", ""),
        ("— OE3: SEGURIDAD BÁSICA —", "", ""),
        ("% portales con SSL válido", pct_true("ssl_ok"), "%"),
        ("% portales con TLS ≥ 1.2", pct_true("tls_aceptable"), "%"),
        ("% portales con HSTS", pct_true("header_hsts"), "%"),
        ("% portales con CSP", pct_true("header_csp"), "%"),
        ("% portales con X-Frame-Options", pct_true("header_x_frame_options"), "%"),
        ("% portales que redirigen HTTP→HTTPS", pct_true("redirige_a_https"), "%"),
        ("% portales que exponen header Server", pct_true("expone_server_header"), "%"),
        ("Score local seguridad medio", media("score_local_security"), "/100"),
    ]
    return pd.DataFrame(filas, columns=["Indicador", "Valor", "Unidad"])


# ---------- API pública ----------

def generar_reporte(df: pd.DataFrame, *, sufijo: str = "") -> Path:
    """
    Genera el archivo Excel y CSV. Devuelve la ruta del .xlsx generado.
    """
    if df.empty:
        raise ValueError("DataFrame de resultados está vacío.")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_base = f"auditoria_egov_suroccidente_{timestamp}{sufijo}"

    # CSV plano (para análisis externo)
    csv_path = REPORTS_DIR / f"{nombre_base}.csv"
    df.to_csv(csv_path, index=False)
    log.info("CSV generado: %s", csv_path)

    # Gráficas
    graficas_dir = REPORTS_DIR / f"graficas_{timestamp}"
    graficas = _generar_graficas(df, graficas_dir)
    log.info("Gráficas generadas: %d", len(graficas))

    # Excel
    wb = Workbook()
    wb.remove(wb.active)

    # H1: Resumen ejecutivo
    ws = wb.create_sheet("1.Resumen Ejecutivo")
    ws["A1"] = "Auditoría Técnica de Portales E-Gov del Suroccidente de Guatemala"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:C1")
    ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:C2")
    resumen_df = _resumen_ejecutivo(df)
    _formatear_hoja(ws, resumen_df, fila_inicio=4)

    # H2: Datos crudos
    ws = wb.create_sheet("2.Datos por Municipalidad")
    # Reducir cols si es muy ancho: tomar las más importantes
    cols_principales = [
        "municipio", "departamento", "url", "reachable", "status_code",
        "https", "ttfb_ms", "tiempo_total_ms", "tamanio_kb",
        "tiene_viewport", "tiene_lang", "imgs_pct_con_alt",
        "score_local_performance",
        "snapshots_unicos", "ultimo_snapshot", "dias_desde_ultima_actualizacion",
        "laip_pct_cumplimiento", "score_local_freshness",
        "ssl_ok", "ssl_tls_version", "ssl_dias_restantes",
        "header_hsts", "header_csp", "redirige_a_https",
        "score_local_security",
    ]
    cols_existentes = [c for c in cols_principales if c in df.columns]
    _formatear_hoja(ws, df[cols_existentes])

    # H3: Descriptiva rendimiento
    ws = wb.create_sheet("3.Descriptiva OE1 Rendim")
    desc1 = A.descriptiva_numerica(df, COLS_PERF)
    _formatear_hoja(ws, desc1)

    # H4: Descriptiva frescura
    ws = wb.create_sheet("4.Descriptiva OE2 Frescura")
    desc2 = A.descriptiva_numerica(df, COLS_FRESHNESS)
    _formatear_hoja(ws, desc2)

    # H5: Descriptiva seguridad
    ws = wb.create_sheet("5.Descriptiva OE3 Segurid")
    desc3 = A.descriptiva_numerica(df, COLS_SECURITY)
    _formatear_hoja(ws, desc3)

    # H6: Proporciones (todas las booleanas)
    ws = wb.create_sheet("6.Proporciones Cumplimnto")
    cols_bool = COLS_BOOL_PERF + COLS_BOOL_FRESHNESS + COLS_BOOL_SECURITY
    props = A.proporciones(df, cols_bool)
    _formatear_hoja(ws, props)

    # H7: Comparación entre departamentos
    ws = wb.create_sheet("7.Comparac Departamentos")
    todas_num = COLS_PERF + COLS_FRESHNESS + COLS_SECURITY
    kw = A.comparar_departamentos_kruskal(df, todas_num)
    if not kw.empty:
        _formatear_hoja(ws, kw)
    else:
        ws["A1"] = "Insuficientes datos para comparar departamentos."
        ws["A1"].font = FONT_BODY

    # H8: Resumen por departamento
    ws = wb.create_sheet("8.Resumen x Departamento")
    rxd = A.resumen_por_departamento(df, todas_num)
    if not rxd.empty:
        _formatear_hoja(ws, rxd)

    # H9: Correlaciones
    ws = wb.create_sheet("9.Correlaciones (Spearman)")
    cols_corr = [
        "tiempo_total_ms", "tamanio_kb", "imgs_pct_con_alt",
        "score_local_performance", "score_local_freshness", "score_local_security",
        "snapshots_unicos", "laip_pct_cumplimiento", "headers_seguridad_pct",
    ]
    corr = A.correlaciones(df, cols_corr)
    if not corr.empty:
        corr_reset = corr.reset_index().rename(columns={"index": "variable"})
        _formatear_hoja(ws, corr_reset)

    # H10: Errores/inalcanzables (diagnóstico)
    ws = wb.create_sheet("10.Errores y No Alcanzados")
    if "reachable" in df.columns:
        no_alc = df[~df["reachable"].fillna(False).astype(bool)][
            [c for c in ["municipio", "departamento", "url", "status_code", "error_fetch"]
             if c in df.columns]
        ]
        if not no_alc.empty:
            _formatear_hoja(ws, no_alc)
        else:
            ws["A1"] = "Todas las URLs auditadas fueron alcanzables."

    # H11 (opcional): Disponibilidad/uptime — solo si venimos del consolidado
    if "uptime_pct" in df.columns:
        ws = wb.create_sheet("11.Disponibilidad (uptime)")
        cols_up = [c for c in ["municipio", "departamento", "url", "n_corridas",
                               "n_exitosas", "uptime_pct", "primera_corrida",
                               "ultima_corrida"] if c in df.columns]
        _formatear_hoja(ws, df[cols_up].sort_values("uptime_pct"))

    # OE4: Análisis de datos categóricos (consume cumple_LAIP / vulnerabilidad)
    try:
        oe4 = A.analisis_oe4_completo(df)
        _agregar_hojas_oe4(wb, oe4)
    except Exception as ex:
        log.exception("Error en análisis OE4: %s", ex)

    xlsx_path = REPORTS_DIR / f"{nombre_base}.xlsx"
    wb.save(xlsx_path)
    log.info("Excel generado: %s", xlsx_path)

    return xlsx_path


def _agregar_hojas_oe4(wb: Workbook, oe4: Dict[str, Any]) -> None:
    """Inserta las hojas del análisis OE4 (χ²/Fisher + regresión logística)."""

    def _hoja_chi2(nombre, titulo, key):
        ws = wb.create_sheet(nombre)
        ws["A1"] = titulo
        ws["A1"].font = FONT_TITLE
        ws.merge_cells("A1:N1")
        tab = oe4.get(key)
        if isinstance(tab, pd.DataFrame) and not tab.empty:
            _formatear_hoja(ws, tab, fila_inicio=3)
        else:
            ws["A3"] = "Datos insuficientes para las pruebas (¿n o variabilidad?)."

    _hoja_chi2("12.OE4 Chi2 LAIP",
               "χ²/Fisher — Cumplimiento LAIP vs predictores categóricos",
               "chi2_laip")
    _hoja_chi2("13.OE4 Chi2 Vulnerab",
               "χ²/Fisher — Vulnerabilidad de seguridad vs predictores",
               "chi2_vuln")

    # Regresiones logísticas
    def _hoja_logit(nombre, titulo, coef_key, met_key):
        ws = wb.create_sheet(nombre)
        ws["A1"] = titulo
        ws["A1"].font = FONT_TITLE
        ws.merge_cells("A1:I1")
        metricas = oe4.get(met_key, {}) or {}
        ws["A2"] = "Métricas del modelo:"
        ws["A2"].font = Font(bold=True)
        fila = 3
        for k, v in metricas.items():
            ws.cell(row=fila, column=1, value=str(k))
            ws.cell(row=fila, column=2, value=_valor_safe(v))
            fila += 1
        fila += 1
        coef = oe4.get(coef_key)
        if isinstance(coef, pd.DataFrame) and not coef.empty:
            ws.cell(row=fila, column=1, value="Coeficientes:").font = Font(bold=True)
            _formatear_hoja(ws, coef, fila_inicio=fila + 1)
        else:
            ws.cell(row=fila, column=1, value="No fue posible ajustar el modelo.")

    _hoja_logit("14.OE4 Regresion LAIP",
                "Regresión logística — Cumplimiento LAIP (positivo = Cumple)",
                "logit_laip_coef", "logit_laip_metricas")
    _hoja_logit("15.OE4 Regresion Vulnerab",
                "Regresión logística — Vulnerabilidad (positivo = Vulnerable)",
                "logit_vuln_coef", "logit_vuln_metricas")

    # Tablas de contingencia LAIP
    ws = wb.create_sheet("16.OE4 Contingencia LAIP")
    ws.cell(row=1, column=1, value="Tablas de contingencia: predictor × Cumplimiento LAIP").font = FONT_TITLE
    fila = 3
    for predictor, tabla in (oe4.get("tablas_contingencia_laip") or {}).items():
        ws.cell(row=fila, column=1, value=predictor).font = Font(bold=True, color="1F4E78")
        fila += 1
        tabla_df = tabla.reset_index()
        _formatear_hoja(ws, tabla_df, fila_inicio=fila)
        fila += len(tabla_df) + 3
